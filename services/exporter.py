import io
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN   = Alignment(vertical="top")
MAX_CELL    = 32000


def safe(val: Any, max_len: int = MAX_CELL) -> str:
    s = str(val or "")
    return s[:max_len] + "...[cut]" if len(s) > max_len else s


def _header_row(ws, headers: List[str]):
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        c = ws.cell(ws.max_row, col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[ws.max_row].height = 28
    ws.freeze_panes = ws.cell(ws.max_row + 1, 1)


def _auto_width(ws, overrides: Dict[str, int] = None):
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        w = min(max(max((len(str(c.value or "")) for c in col), default=0) + 2, 10), 60)
        if overrides and cl in overrides:
            w = overrides[cl]
        ws.column_dimensions[cl].width = w


def build_excel(evaluations: List[Dict], messages: List[Dict]) -> bytes:
    """
    evaluations: list of dicts from the DB (joined ticket + evaluation)
    messages:    list of {ticket_id, role, ts, body} dicts
    Returns: bytes of the .xlsx file
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Tickets + QA ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tickets & QA"
    h1 = [
        "Ticket ID", "Subject", "Agent", "Group", "CSAT", "Created", "Resolved",
        "Total Score", "Complexity", "Churn Risk", "Churn Reason", "Contact Problem",
        "Summary", "Coaching Tip",
        "Clarity", "Tone", "Empathy", "Accuracy", "Resolution",
        "Efficiency", "Ownership", "Commercial",
    ]
    _header_row(ws1, h1)

    for ev in evaluations:
        sc = ev.get("scores") or {}
        ws1.append([
            ev.get("ticket_id"),
            safe(ev.get("subject"), 300),
            safe(ev.get("agent_name"), 100),
            safe(ev.get("group_name"), 100),
            ev.get("csat"),
            str(ev.get("created_at") or "")[:10],
            str(ev.get("resolved_at") or "")[:10],
            ev.get("total_score"),
            ev.get("complexity"),
            "YES" if ev.get("churn_risk_flag") else "",
            safe(ev.get("churn_risk_reason"), 200),
            "YES" if ev.get("contact_problem_flag") else "",
            safe(ev.get("summary"), 500),
            safe(ev.get("coaching_tip"), 300),
            (sc.get("clarity_structure") or sc.get("clarity") or {}).get("score", ""),
            (sc.get("tone_professionalism") or sc.get("tone") or {}).get("score", ""),
            (sc.get("empathy") or {}).get("score", ""),
            (sc.get("accuracy") or {}).get("score", ""),
            (sc.get("resolution_quality") or {}).get("score", ""),
            (sc.get("efficiency") or {}).get("score", ""),
            (sc.get("ownership") or {}).get("score", ""),
            (sc.get("commercial_awareness") or {}).get("score", ""),
        ])
        for col in range(1, len(h1) + 1):
            c = ws1.cell(ws1.max_row, col)
            c.font = BODY_FONT
            c.alignment = WRAP_ALIGN if col in (13, 14) else TOP_ALIGN
    _auto_width(ws1, {"B": 50, "K": 40, "M": 60, "N": 50})

    # ── Sheet 2: Full Conversations ──────────────────────────────────────────
    ws2 = wb.create_sheet("Conversations")
    h2 = ["Ticket ID", "Subject", "Agent", "Churn Risk", "Msg #", "Role", "Timestamp", "Message"]
    _header_row(ws2, h2)

    # Build subject/agent lookup from evaluations
    meta = {ev["ticket_id"]: ev for ev in evaluations}
    msg_counter: Dict[str, int] = {}

    for msg in messages:
        tid = msg["ticket_id"]
        msg_counter[tid] = msg_counter.get(tid, 0) + 1
        ev = meta.get(tid, {})
        ws2.append([
            tid,
            safe(ev.get("subject"), 200),
            safe(ev.get("agent_name"), 100),
            "YES" if ev.get("churn_risk_flag") else "",
            msg_counter[tid],
            msg["role"],
            str(msg.get("ts") or "")[:16].replace("T", " "),
            safe(msg["body"]),
        ])
        for col in range(1, len(h2) + 1):
            c = ws2.cell(ws2.max_row, col)
            c.font = BODY_FONT
            c.alignment = WRAP_ALIGN if col == 8 else TOP_ALIGN
    _auto_width(ws2, {"B": 45, "H": 120})

    # ── Sheet 3: Agent Summary ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Agent Summary")
    _header_row(ws3, [
        "Agent", "Group", "Tickets", "Avg QA Score", "QA Risk Index",
        "Churn Flags", "Churn %", "Contact Problem %", "Tier",
    ])
    ag: Dict[str, Dict] = {}
    for ev in evaluations:
        k = ev.get("agent_name") or "Unknown"
        if k not in ag:
            ag[k] = {"group": ev.get("group_name", ""), "evals": [], "churn": 0, "contact": 0}
        ag[k]["evals"].append(ev)
        if ev.get("churn_risk_flag"):
            ag[k]["churn"] += 1
        if ev.get("contact_problem_flag"):
            ag[k]["contact"] += 1

    for name, a in sorted(ag.items(), key=lambda x: -len(x[1]["evals"])):
        n = len(a["evals"])
        avg = round(sum(e.get("total_score") or 0 for e in a["evals"]) / n, 1)
        risk = round(sum((5 - (e.get("total_score") or 0) / 20) * (e.get("arr") or 0) for e in a["evals"]))
        ws3.append([
            name, a["group"], n, avg, risk,
            a["churn"], f"{round(a['churn'] / n * 100)}%",
            f"{round(a['contact'] / n * 100)}%",
            "🟢 Top" if avg >= 75 and a["churn"] / n <= 0.1
            else "🟡 Solid" if avg >= 65
            else "🟠 Needs Improvement" if avg >= 55
            else "🔴 At Risk",
        ])
        for col in range(1, 10):
            ws3.cell(ws3.max_row, col).font = BODY_FONT
    _auto_width(ws3, {"A": 24, "B": 20})

    # ── Sheet 4: Churn Risk ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("Churn Risk")
    _header_row(ws4, [
        "Ticket ID", "Subject", "Agent", "Group", "Score",
        "CSAT", "Churn Reason", "Coaching Tip",
    ])
    churn_evals = [ev for ev in evaluations if ev.get("churn_risk_flag")]
    if churn_evals:
        for ev in churn_evals:
            ws4.append([
                ev.get("ticket_id"),
                safe(ev.get("subject"), 300),
                safe(ev.get("agent_name"), 100),
                safe(ev.get("group_name"), 100),
                ev.get("total_score"),
                ev.get("csat"),
                safe(ev.get("churn_risk_reason"), 300),
                safe(ev.get("coaching_tip"), 300),
            ])
            for col in range(1, 9):
                ws4.cell(ws4.max_row, col).font = BODY_FONT
    else:
        ws4.append(["No churn risk tickets in this export."])
    _auto_width(ws4, {"B": 50, "G": 60, "H": 60})

    # ── Sheet 5: Score Evidence ──────────────────────────────────────────────
    ws5 = wb.create_sheet("Score Evidence")
    _header_row(ws5, ["Ticket ID", "Agent", "Category", "Score", "Evidence"])
    cats = [
        "clarity_structure", "tone_professionalism", "empathy", "accuracy",
        "resolution_quality", "efficiency", "ownership", "commercial_awareness",
    ]
    for ev in evaluations:
        sc = ev.get("scores") or {}
        for cat in cats:
            s = sc.get(cat) or {}
            score = s.get("score")
            if score is not None:
                ws5.append([
                    ev.get("ticket_id"),
                    safe(ev.get("agent_name"), 100),
                    cat.replace("_", " "),
                    score,
                    safe(s.get("reason", ""), 300),
                ])
                for col in range(1, 6):
                    ws5.cell(ws5.max_row, col).font = BODY_FONT
    _auto_width(ws5, {"C": 24, "E": 100})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
